#!/usr/bin/env python3
"""Scan procurement-list workbooks for blank specifications."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADERS = ["所属分类", "设备名称", "技术参数/规格", "数量", "单位", "可补充候选参数", "核查备注"]


def norm(value) -> str:
    return str(value).strip() if value is not None else ""


def read_procurement_rows(xlsx: Path) -> list[dict]:
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb.active
    rows: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        vals = [norm(v) for v in row[:5]]
        if len(vals) < 5 or not any(vals):
            continue
        if vals[1] in {"设备名称", "设备材料名称"}:
            continue
        rows.append({"name": vals[1], "spec": vals[2], "qty": vals[3], "unit": vals[4]})
    return rows


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9D9D9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [18, 28, 28, 12, 10, 40, 44]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def main() -> int:
    parser = argparse.ArgumentParser(description="Create 未明确确认部件规格参数.xlsx from a procurement folder.")
    parser.add_argument("procurement_folder", type=Path)
    parser.add_argument("--out", type=Path, default=None)
    args = parser.parse_args()

    root = args.procurement_folder.resolve()
    out = args.out.resolve() if args.out else root.parent / "未明确确认部件规格参数.xlsx"

    records: list[list[str]] = []
    for xlsx in sorted(root.rglob("*.xlsx")):
        if xlsx.name.startswith("~$"):
            continue
        category = xlsx.parent.name
        for row in read_procurement_rows(xlsx):
            if row["name"] and not row["spec"]:
                records.append([category, row["name"], "", row["qty"], row["unit"], "", "规格参数为空，需结合设计文件、说明书或厂家深化资料确认。"])

    wb = Workbook()
    ws = wb.active
    ws.title = "未明确确认部件规格参数"
    ws.append(HEADERS)
    for record in records:
        ws.append(record)
    style_sheet(ws)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"saved={out}")
    print(f"blank_spec_rows={len(records)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
